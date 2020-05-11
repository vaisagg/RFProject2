import openpyxl


def ReadTD_Rowcount(sheetName):
    wb = openpyxl.load_workbook("..//TestData//TestData.xlsx")
    sh = wb[sheetName]
    return sh.max_row


def ReadTD(sheetName,rowval,colval):
    wb = openpyxl.load_workbook("..//TestData//TestData.xlsx")
    sh = wb[sheetName]
    cellval = sh.cell(int(rowval),int(colval))
    return cellval.value




